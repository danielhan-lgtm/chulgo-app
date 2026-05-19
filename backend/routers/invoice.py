import io
import json
import os
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime
from typing import List, Optional

import requests
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from pydantic import BaseModel

from utils_core import BASE_URL, api_headers, fetch_all_items_list, load_config

router = APIRouter()

CUSTOMERS_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "customers.json")
TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "templates", "거래명세서_양식.xlsx")


# ── 거래처 관리 ───────────────────────────────────────────────────────────────

class Customer(BaseModel):
    id: Optional[str] = None
    name: str
    business_no: str = ""
    representative: str = ""
    address: str = ""
    phone: str = ""
    email: str = ""


def _load_customers() -> list:
    if not os.path.exists(CUSTOMERS_PATH):
        return []
    with open(CUSTOMERS_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def _save_customers(data: list):
    with open(CUSTOMERS_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


@router.get("/customers")
def list_customers():
    return _load_customers()


@router.post("/customers")
def create_customer(body: Customer):
    customers = _load_customers()
    body.id = str(uuid.uuid4())
    customers.append(body.dict())
    _save_customers(customers)
    return body


@router.put("/customers/{cid}")
def update_customer(cid: str, body: Customer):
    customers = _load_customers()
    for i, c in enumerate(customers):
        if c["id"] == cid:
            body.id = cid
            customers[i] = body.dict()
            _save_customers(customers)
            return body
    raise HTTPException(status_code=404, detail="거래처를 찾을 수 없습니다.")


@router.delete("/customers/{cid}")
def delete_customer(cid: str):
    customers = _load_customers()
    customers = [c for c in customers if c["id"] != cid]
    _save_customers(customers)
    return {"ok": True}


# ── BoxHero 판매 내역 조회 ──────────────────────────────────────────────────

@router.get("/sales")
def get_sales(from_date: str, to_date: str, location_id: Optional[int] = None):
    cfg = load_config()
    token = cfg.get("api_token", "")
    if not token:
        raise HTTPException(status_code=400, detail="BoxHero API 토큰이 설정되어 있지 않습니다. 설정 페이지에서 토큰을 입력해주세요.")

    # 아이템 id → sku 매핑 캐시 (상세 응답에 name은 있지만 sku는 없음)
    all_items = fetch_all_items_list(token)
    item_sku_map = {item["id"]: item.get("sku", "") for item in all_items}

    # 날짜 필터 파싱
    try:
        dt_from = datetime.strptime(from_date, "%Y-%m-%d").date()
        dt_to = datetime.strptime(to_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="날짜 형식이 올바르지 않습니다. YYYY-MM-DD")

    # 출고 내역 목록 조회 (페이지네이션) — 요약 정보만 포함
    summaries = []
    cursor = None
    while True:
        params: dict = {"type": "out", "limit": 100}
        if cursor:
            params["cursor"] = cursor
        if location_id:
            params["location_id"] = location_id
        try:
            r = requests.get(f"{BASE_URL}/v1/location-txs", headers=api_headers(token), params=params, timeout=15)
            r.raise_for_status()
        except requests.HTTPError as e:
            raise HTTPException(status_code=502, detail=f"BoxHero API 오류: {e}")
        data = r.json()
        for tx in data.get("items", []):
            tx_date_str = (tx.get("transaction_time") or tx.get("created_at") or "")[:10]
            try:
                tx_date = datetime.strptime(tx_date_str, "%Y-%m-%d").date()
            except ValueError:
                continue
            if dt_from <= tx_date <= dt_to:
                summaries.append({"id": tx["id"], "date": tx_date_str, "memo": tx.get("memo", "")})
        # 커서가 가장 오래된 항목 기준 — 기간 이전 항목이 나오면 중단
        if not data.get("has_more"):
            break
        cursor = data.get("cursor")

    # 각 거래 상세 병렬 호출 — GET /v1/location-txs/{id}
    def fetch_detail(summary: dict) -> dict:
        try:
            r = requests.get(
                f"{BASE_URL}/v1/location-txs/{summary['id']}",
                headers=api_headers(token),
                timeout=15,
            )
            r.raise_for_status()
            detail = r.json().get("item", {})
        except Exception:
            detail = {}

        items_in_tx = []
        for row in detail.get("items", []):
            if row.get("deleted"):
                continue
            item_id = row.get("id")
            qty = abs(row.get("quantity", 0))
            items_in_tx.append({
                "item_id": item_id,
                "item_name": row.get("name", f"아이템({item_id})"),
                "sku": item_sku_map.get(item_id, ""),
                "qty": qty,
                "unit_price": 0,
                "remark": "",
            })

        return {
            "tx_id": summary["id"],
            "date": summary["date"],
            "memo": summary["memo"],
            "items": items_in_tx,
        }

    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = {pool.submit(fetch_detail, s): s for s in summaries}
        results = {s["id"]: None for s in summaries}
        for future in as_completed(futures):
            s = futures[future]
            results[s["id"]] = future.result()

    # 날짜순 정렬
    filtered = sorted(
        [v for v in results.values() if v],
        key=lambda x: x["date"],
        reverse=True,
    )
    return {"transactions": filtered}


@router.get("/sales/debug")
def debug_sales_raw():
    """BoxHero 실제 응답 구조 확인용 — 첫 거래 1건 원본 반환"""
    cfg = load_config()
    token = cfg.get("api_token", "")
    if not token:
        raise HTTPException(status_code=400, detail="토큰 없음")
    r = requests.get(
        f"{BASE_URL}/v1/location-txs",
        headers=api_headers(token),
        params={"type": "out", "limit": 1},
        timeout=15,
    )
    r.raise_for_status()
    return r.json()


# ── 템플릿 경로 조회/변경 ──────────────────────────────────────────────────

@router.get("/template-path")
def get_template_path():
    cfg = load_config()
    custom = cfg.get("invoice_template_path", "")
    path = custom if custom and os.path.exists(custom) else TEMPLATE_PATH
    return {"path": path, "exists": os.path.exists(path)}


class TemplatePathRequest(BaseModel):
    path: str


@router.post("/template-path")
def set_template_path(body: TemplatePathRequest):
    if not os.path.exists(body.path):
        raise HTTPException(status_code=400, detail=f"파일을 찾을 수 없습니다: {body.path}")
    from utils_core import save_config
    save_config({"invoice_template_path": body.path})
    return {"ok": True, "path": body.path}


# ── 거래명세서 엑셀 생성 ────────────────────────────────────────────────────

class InvoiceItem(BaseModel):
    item_name: str
    sku: str = ""
    qty: int
    unit_price: int
    remark: str = ""


class GenerateRequest(BaseModel):
    customer_id: str
    issue_date: str           # 거래일자
    doc_number: str = ""      # 관리번호
    trade_name: str = ""      # 거래건명
    payment_terms: str = ""   # 결제조건
    items: List[InvoiceItem]


@router.post("/generate")
def generate_invoice(body: GenerateRequest):
    import traceback
    try:
        # 거래처 조회
        customers = _load_customers()
        customer = next((c for c in customers if c["id"] == body.customer_id), None)
        if not customer:
            raise HTTPException(status_code=404, detail="거래처를 찾을 수 없습니다.")

        cfg = load_config()
        tpl = cfg.get("invoice_template_path", "") or TEMPLATE_PATH
        if not os.path.exists(tpl):
            raise HTTPException(status_code=500, detail=f"거래명세서 양식 파일을 찾을 수 없습니다: {tpl}")

        # 7개씩 청크 분할 (7개 초과 시 여러 시트 생성)
        PAGE_SIZE = 7
        chunks = [body.items[i:i + PAGE_SIZE] for i in range(0, len(body.items), PAGE_SIZE)]
        total_pages = len(chunks)

        wb = load_workbook(tpl)
        template_ws = wb.worksheets[0]

        # 추가 페이지가 필요한 경우 미리 복사 (원본 수정 전에 복사해야 함)
        extra_sheets = []
        for _ in range(1, total_pages):
            extra_sheets.append(wb.copy_worksheet(template_ws))

        def fill_sheet(ws, chunk: list, page: int):
            if body.doc_number:
                doc_no = body.doc_number if page == 1 else f"{body.doc_number}-{page - 1}"
            else:
                doc_no = "" if page == 1 else f"-{page - 1}"
            ws["B2"] = f"관리번호:  {doc_no}" if doc_no else "관리번호:"
            ws["E8"] = customer.get("name", "")
            ws["E9"] = body.trade_name
            ws["E10"] = body.issue_date
            ws["E11"] = body.payment_terms
            for idx, item in enumerate(chunk):
                row = 16 + idx
                ws[f"B{row}"] = item.item_name
                ws[f"I{row}"] = item.qty
                ws[f"K{row}"] = item.unit_price
                if item.remark:
                    ws[f"Q{row}"] = item.remark

        fill_sheet(template_ws, chunks[0], 1)
        for i, (chunk, ws) in enumerate(zip(chunks[1:], extra_sheets), 2):
            fill_sheet(ws, chunk, i)

        # 메모리 버퍼로 저장
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        customer_name = customer.get("name", "거래처")
        filename = f"거래명세서_{customer_name}_{body.issue_date}.xlsx"

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{requests.utils.quote(filename)}"},
        )
    except HTTPException:
        raise
    except Exception as e:
        tb = traceback.format_exc()
        raise HTTPException(status_code=500, detail=f"엑셀 생성 오류: {type(e).__name__}: {e}\n{tb}")
